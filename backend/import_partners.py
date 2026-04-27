"""
Import partners and vehicles from Excel file into the database.
Run this script to populate the partners and vehicles tables.

Usage:
    python import_partners.py /path/to/RentACar_Partners_Consolidated.xlsx
"""

import sys
from datetime import datetime, timezone

# Add openpyxl to requirements if not already
try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from models import Base, PartnerDB, VehicleDB


def parse_price(value) -> float | None:
    """Parse price from various formats like '€77.39', '77.39', or numeric."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remove currency symbols, spaces, and handle European number format
        cleaned = value.replace('€', '').replace(' ', '').strip()
        if not cleaned:
            return None
        # Handle European format (1.234,56 -> 1234.56)
        if ',' in cleaned and '.' in cleaned:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif ',' in cleaned:
            cleaned = cleaned.replace(',', '.')
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def parse_franchise(value) -> float | None:
    """Parse franchise value, handling special cases like '€0.00' or '€1,200.00'."""
    return parse_price(value)


def parse_int(value) -> int | None:
    """Parse integer value."""
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        try:
            return int(float(value))
        except ValueError:
            return None
    return None


def import_from_excel(excel_path: str, database_url: str = "sqlite:///data/app.db"):
    """Import partners and vehicles from Excel file."""

    # Create database connection
    engine = create_engine(database_url)
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Load Excel file
    print(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    # Get the first sheet (or you can specify sheet name)
    ws = wb.active
    print(f"Reading sheet: {ws.title}")

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")

    # Map headers to expected columns
    header_map = {}
    for idx, header in enumerate(headers):
        if header:
            header_lower = header.lower().strip()
            if 'partner' in header_lower:
                header_map['partner'] = idx
            elif 'category' in header_lower or 'categoria' in header_lower:
                header_map['category'] = idx
            elif 'group' in header_lower or 'code' in header_lower or 'grupo' in header_lower:
                header_map['group_code'] = idx
            elif 'vehicle' in header_lower or 'description' in header_lower or 'descri' in header_lower:
                header_map['description'] = idx
            elif '1-3' in header_lower or '(1-3' in header_lower:
                header_map['price_1_3'] = idx
            elif '4-6' in header_lower or '(4-6' in header_lower:
                header_map['price_4_6'] = idx
            elif '7-29' in header_lower or '(7-29' in header_lower:
                header_map['price_7_29'] = idx
            elif 'monthly' in header_lower or 'mensal' in header_lower:
                header_map['price_monthly'] = idx
            elif 'franchise' in header_lower or 'franquia' in header_lower:
                header_map['franchise'] = idx
            elif 'min age' in header_lower or 'idade' in header_lower:
                header_map['min_age'] = idx
            elif 'license' in header_lower or 'carta' in header_lower:
                header_map['license_years'] = idx
            elif 'note' in header_lower or 'obs' in header_lower:
                header_map['notes'] = idx

    print(f"Column mapping: {header_map}")

    # Track partners we've created
    partners_cache = {}
    vehicles_added = 0

    # Process each row (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Get partner name
        partner_name = row[header_map.get('partner', 0)] if header_map.get('partner') is not None else None
        if not partner_name:
            continue

        partner_name = str(partner_name).strip()

        # Get or create partner
        if partner_name not in partners_cache:
            # Check if partner exists in DB
            existing = session.query(PartnerDB).filter(PartnerDB.name == partner_name).first()
            if existing:
                partners_cache[partner_name] = existing
                print(f"Found existing partner: {partner_name}")
            else:
                # Create new partner
                new_partner = PartnerDB(
                    name=partner_name,
                    is_active=1,
                    created_at=datetime.now(timezone.utc),
                    updated_at=datetime.now(timezone.utc)
                )
                session.add(new_partner)
                session.flush()  # Get the ID
                partners_cache[partner_name] = new_partner
                print(f"Created new partner: {partner_name} (ID: {new_partner.id})")

        partner = partners_cache[partner_name]

        # Get vehicle data
        category = row[header_map.get('category', 1)] if header_map.get('category') is not None else None
        group_code = row[header_map.get('group_code', 2)] if header_map.get('group_code') is not None else None
        description = row[header_map.get('description', 3)] if header_map.get('description') is not None else None

        if not category or not group_code or not description:
            print(f"Row {row_idx}: Skipping - missing required fields (category/group_code/description)")
            continue

        # Parse prices
        price_1_3 = parse_price(row[header_map['price_1_3']]) if header_map.get('price_1_3') is not None else None
        price_4_6 = parse_price(row[header_map['price_4_6']]) if header_map.get('price_4_6') is not None else None
        price_7_29 = parse_price(row[header_map['price_7_29']]) if header_map.get('price_7_29') is not None else None
        price_monthly = parse_price(row[header_map['price_monthly']]) if header_map.get('price_monthly') is not None else None
        franchise = parse_franchise(row[header_map['franchise']]) if header_map.get('franchise') is not None else None
        min_age = parse_int(row[header_map['min_age']]) if header_map.get('min_age') is not None else None
        license_years = parse_int(row[header_map['license_years']]) if header_map.get('license_years') is not None else None
        notes = str(row[header_map['notes']]).strip() if header_map.get('notes') is not None and row[header_map['notes']] else None

        # Check if vehicle already exists
        existing_vehicle = session.query(VehicleDB).filter(
            VehicleDB.partner_id == partner.id,
            VehicleDB.group_code == str(group_code).strip()
        ).first()

        if existing_vehicle:
            # Update existing vehicle
            existing_vehicle.category = str(category).strip()
            existing_vehicle.description = str(description).strip()
            existing_vehicle.price_1_3_days = price_1_3
            existing_vehicle.price_4_6_days = price_4_6
            existing_vehicle.price_7_29_days = price_7_29
            existing_vehicle.price_monthly = price_monthly
            existing_vehicle.franchise = franchise
            existing_vehicle.min_age = min_age
            existing_vehicle.license_years = license_years
            existing_vehicle.notes = notes
            existing_vehicle.updated_at = datetime.now(timezone.utc)
            print(f"  Updated vehicle: {group_code} - {description[:30]}...")
        else:
            # Create new vehicle
            new_vehicle = VehicleDB(
                partner_id=partner.id,
                category=str(category).strip(),
                group_code=str(group_code).strip(),
                description=str(description).strip(),
                price_1_3_days=price_1_3,
                price_4_6_days=price_4_6,
                price_7_29_days=price_7_29,
                price_monthly=price_monthly,
                franchise=franchise,
                min_age=min_age,
                license_years=license_years,
                notes=notes,
                is_active=1,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )
            session.add(new_vehicle)
            vehicles_added += 1
            print(f"  Added vehicle: {group_code} - {description[:30]}...")

    # Commit all changes
    session.commit()

    print("\n" + "="*50)
    print(f"Import completed!")
    print(f"Partners: {len(partners_cache)}")
    print(f"Vehicles added: {vehicles_added}")
    print("="*50)

    # Show summary
    print("\nPartners in database:")
    for partner in session.query(PartnerDB).filter(PartnerDB.is_active == 1).all():
        vehicle_count = session.query(VehicleDB).filter(VehicleDB.partner_id == partner.id).count()
        print(f"  - {partner.name}: {vehicle_count} vehicles")

    session.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_partners.py /path/to/excel_file.xlsx")
        print("       python import_partners.py /path/to/excel_file.xlsx sqlite:///data/app.db")
        sys.exit(1)

    excel_path = sys.argv[1]
    database_url = sys.argv[2] if len(sys.argv) > 2 else "sqlite:///data/app.db"

    import_from_excel(excel_path, database_url)
